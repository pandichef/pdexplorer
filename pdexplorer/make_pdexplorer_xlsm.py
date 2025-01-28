"""
saves python in cells as strings.
you have to manually remove the leading apostrophe in Excel to get the formulas to work.
"""
import os
import shutil
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

file_dir = os.path.dirname(os.path.abspath(__file__))

def python_to_vba_string(python_code, module_name):
    """
    Converts Python code into a VBA-compatible string for embedding in Excel.

    Args:
        python_code (str): The Python code to convert.
        module_name (str): The name of the Python module to append.

    Returns:
        str: The VBA-compatible string representation.
    """
    # Escape double quotes for VBA
    python_code = python_code.replace('"', '""""')

    # Split the Python code into lines and format them for VBA
    vba_lines = []
    for line in python_code.splitlines():
        if line.strip():  # Non-empty line
            vba_lines.append(f'"{line}" & Chr(10)')
        else:  # Empty line
            vba_lines.append('" & Chr(10)')

    # Add the module name and ensure VBA concatenation
    vba_lines.append(f'"""{module_name}""" & Chr(10)')

    # Join the lines, adding VBA concatenation (`& _`) for long lines
    vba_string = " & _\n        ".join(vba_lines)

    # Wrap everything in =PY(...) as required
    return f"=PY({vba_string})"

# Example usage
# if __name__ == "__main__":
#     python_code = '''
# a = "hello"
# b = "world"
# a + " " + b
# '''
#     module_name = "tmp123.py"
#     vba_code = python_to_vba_string(python_code, module_name)
#     print(vba_code)


def make_xlsm_file_with_py_function_hack(filename, columnA_contents):
    """
    openpyxl can't directly create a =py() function
    Modern Excel forces the function to be =@py(), which errors out
    """
    if not filename.lower().endswith('.xlsm'):
        raise ValueError(f"Only .xlsm files are supported. Provided: '{filename}'")

    # If the file does not exist, create an empty macro-enabled workbook
    if not os.path.exists(filename):
        template_path = os.path.join(file_dir,"_template.xlsm")  # Path to a macro-enabled workbook template
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template '{template_path}' required to create .xlsm files is missing.")
        shutil.copy(template_path, filename)
        print(f"Created new macro-enabled workbook '{filename}' using the template.")

    # Now load with keep_vba to preserve macros
    try:
        wb = load_workbook(filename, keep_vba=True)
    except Exception as e:
        raise OSError(f"Unable to load '{filename}' as a macro-enabled workbook.") from e

    # Either clear or create the _pdexplorer sheet
    if "_pdexplorer" in wb.sheetnames:
        ws = wb["_pdexplorer"]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet("_pdexplorer")


    # Write the DataFrame to the _pdexplorer sheet
    df = pd.DataFrame({"A": columnA_contents})
    for i, value in enumerate(df["A"], start=1):
        if isinstance(value, str) and value.startswith("=PY("):
            ws[f"A{i}"].value = value  # '=PY("a = 1")'  # Assign the formula
            ws[f"A{i}"].data_type = "s"  # Explicitly mark as string type
        else:
            ws[f"A{i}"].value = value

    # Save the workbook
    wb.save(filename)
    wb.close()
    print(f"Sheet '_pdexplorer' updated in '{filename}'.")


def make_module_contents_list(module_names):
    """
    Given a list of module filenames (like ["_search.py"]), read each file's contents
    from the current directory, remove full-line comments (prefixed with "#"),
    then wrap them in a formula string =PY("...").
    Return a list of such strings.
    """
    import os
    module_formulas = []
    for module_name in module_names:
        # with open(os.path.join('pdexplorer', module_name), "r", encoding="utf-8") as f:
        # print(os.getcwd())
        # print(os.getcwd())
        # print(os.getcwd())
        with open(os.path.join(file_dir, module_name), "r", encoding="utf-8") as f:
            lines = f.readlines()

        processed_lines = []
        for line in lines:
            stripped_line = line.strip()
           
            if stripped_line.startswith("#") \
                or stripped_line.startswith("from .") \
                or stripped_line.startswith("from xlwings"):
                continue # line to skip
            if stripped_line.startswith("from rich import print as rich_print"):
                processed_lines.append("rich_print = print\n")
            else:
                processed_lines.append(line)

        # Join the processed lines
        filtered_code = "".join(processed_lines)

        # Add module name and wrap in =PY
        filtered_code += f'\n\n"{module_name}"'
        formula_str = f"""'=PY({filtered_code}\n"""
        module_formulas.append(formula_str)

    return module_formulas

# def fix_py_function_perhaps_using_vba(xlsm_path):
#     """
#     Injects a VBA module 'Module_PyFix' into the .xlsm, then runs a macro that:
#       1) Finds cells in _pdexplorer whose text begins with '=PY(...),
#          removes the leading apostrophe so it's recognized as a formula
#       2) Simulates Ctrl+Enter in each such cell (SendKeys).

#     Requires:
#     - Windows + Excel + pywin32
#     - "Trust access to the VBA project object model" enabled in Excel Options > Trust Center
#     - Excel must be able to accept =PY(...) formulas (Insider build).
#     """
#     import os
#     import win32com.client as win32

#     if not os.path.exists(xlsm_path):
#         raise FileNotFoundError(f"File not found: '{xlsm_path}'")

#     # Launch Excel via COM, make it visible so SendKeys actually works
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     excel.Visible = True  # must be TRUE for SendKeys to work reliably
#     wb = excel.Workbooks.Open(os.path.abspath(xlsm_path))

#     # Get the VBA project
#     vbproject = wb.VBProject

#     # Create or get a module named "Module_PyFix"
#     module_name = "Module_PyFix"
#     try:
#         vb_module = vbproject.VBComponents(module_name)
#     except:
#         # 1 = vbext_ct_StdModule (use integer constant to avoid makepy issues)
#         vb_module = vbproject.VBComponents.Add(1)
#         vb_module.Name = module_name

#     # VBA code that:
#     #   1) Activates the _pdexplorer sheet
#     #   2) For each cell, if it starts with ' =PY(...), remove leading apostrophe
#     #   3) Press Ctrl+Enter (SendKeys "^({ENTER})") to re‐enter the formula
#     vba_code = r'''
# Public Sub FixPyFormulasAndReEnter()
#     On Error Resume Next
#     Dim ws As Worksheet
#     Set ws = ThisWorkbook.Worksheets("_pdexplorer")
#     If ws Is Nothing Then Exit Sub
    
#     ' Ensure Excel is visible so SendKeys works
#     Application.Visible = True
#     ws.Activate

#     Dim c As Range
#     For Each c In ws.UsedRange
#         If VarType(c.Value) = vbString Then
#             If Left(c.Value, 5) = "'=PY(" Then
#                 ' Select the cell so SendKeys targets it
#                 c.Select
#                 ' Remove the apostrophe so we get a real formula
#                 c.Formula = Mid(c.Value, 2)
#                 ' Force Excel to "re-enter" via Ctrl+Enter
#                 Application.SendKeys "^({ENTER})", True
#                 DoEvents
#             End If
#         End If
#     Next c
# End Sub
# '''

#     # Insert or replace the code in the module
#     code_module = vb_module.CodeModule
#     code_module.DeleteLines(1, code_module.CountOfLines)
#     code_module.AddFromString(vba_code)

#     # Run the macro
#     macro_name = f"{module_name}.FixPyFormulasAndReEnter"
#     excel.Run(macro_name)

#     # Save & close
#     wb.Save()
#     wb.Close()
#     excel.Quit()

#     print("VBA macro injected, executed, and the workbook saved.")


# def fix_py_function_perhaps_using_vba(xlsm_path):
#     """
#     Injects a VBA module into the .xlsm and runs a macro that:
#     1) Finds cells in _pdexplorer that start with "'=PY(".
#     2) Removes the apostrophe so the cell has a real "=PY(...)" formula.
#     3) Presses Ctrl+Enter via SendKeys to re-enter the formula.
    
#     Updated to set cell.Formula2R1C1 (rather than cell.Formula) 
#     for improved compatibility with dynamic (Python) formulas.
#     """
#     import os
#     import win32com.client as win32

#     if not os.path.exists(xlsm_path):
#         raise FileNotFoundError(f"File not found: '{xlsm_path}'")

#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     excel.Visible = True  # Must be visible if using SendKeys
#     wb = excel.Workbooks.Open(os.path.abspath(xlsm_path))

#     vbproject = wb.VBProject
#     module_name = "Module_PyFix"
#     try:
#         vb_module = vbproject.VBComponents(module_name)
#     except:
#         vb_module = vbproject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
#         vb_module.Name = module_name

#     # Updated macro uses Formula2R1C1
#     vba_code = r'''
# Option Explicit

# Public Sub FixPyFormulasAndReEnter()
#     On Error Resume Next
#     Dim ws As Worksheet
#     Set ws = ThisWorkbook.Worksheets("_pdexplorer")
#     If ws Is Nothing Then Exit Sub
    
#     Application.Visible = True
#     ws.Activate

#     Dim c As Range
#     For Each c In ws.UsedRange
#         If VarType(c.Value) = vbString Then
#             If Left(c.Value, 5) = "'=PY(" Then
#                 c.Select
#                 ' Use .Formula2R1C1 to handle Python dynamic array formulas
#                 c.Formula2R1C1 = Mid(c.Value, 2)
#                 ' Now press Ctrl+Enter to "re-enter" the formula
#                 Application.SendKeys "^({ENTER})", True
#                 DoEvents
#             End If
#         End If
#     Next c
# End Sub
# '''

#     code_module = vb_module.CodeModule
#     code_module.DeleteLines(1, code_module.CountOfLines)
#     code_module.AddFromString(vba_code)

#     # Run the macro
#     excel.Run(f"{module_name}.FixPyFormulasAndReEnter")

#     # Save & close
#     wb.Save()
#     wb.Close()
#     excel.Quit()

#     print("VBA macro injected, executed, and the workbook saved.")


# if __name__ == "__main__":
def main():
    try:
        filename = sys.argv[1]
    except IndexError:
        raise Exception("Please provide a filename as an argument.")
    module_names = ["_search.py","_dataset.py","_get_custom_attributes.py","_patsify.py","_print.py",
                    "_commandarg.py","_quietly.py","_print_horizontal_line.py","preserve.py","use.py",
                    "sort.py","_by.py","regress.py","scatter.py","histogram.py","logit.py","drop.py"]
    # module_names = ['tmp123.py']
    columnA_contents = make_module_contents_list(module_names)
    make_xlsm_file_with_py_function_hack(filename, columnA_contents)
#     fix_py_function_perhaps_using_vba("example3.xlsm")
#     print(python_to_vba_string("""
# a = "hello"
# b = "world"
# a + " " + b
# """, "tmp123.py"))

if __name__ == "__main__":
    main()
